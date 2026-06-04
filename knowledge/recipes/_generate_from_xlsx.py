"""엑셀 목록표(aim_odysseus-추가를위한목록표.xlsx) → 레시피 YAML 시드 생성기.

사용:
    python knowledge/recipes/_generate_from_xlsx.py [엑셀경로]

엑셀 구조: A=No, B=생성물(산출물), C~F=inputA~D(입력 문서 라벨).
B의 산출물 1행 = 레시피 1개(YAML) = 스튜디오 버튼 1개.

※ 기존 같은 id 파일이 있으면 덮어쓰지 않는다(사용자가 손본 내용 보호).
   다시 만들고 싶으면 해당 yaml 을 지우고 실행.
"""
from __future__ import annotations

import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(
    HERE, "..", "..", "..", "_context", "aim_odysseus-추가를위한목록표.xlsx"
)

# 출력명 → 기본 포맷 추론 (확장자는 "요청중" 이라 합리적 기본값 사용; 추후 yaml 수정)
EXT_FORMAT = {"pptx": "pptx", "ppt": "pptx", "xlsx": "xlsx", "hwpx": "hwpx",
              "hwp": "hwpx", "docx": "docx", "pdf": "hwpx"}

CATEGORY_RULES = [
    ("발표자료", ["wrap up", "presentation", "발표", "ppt"]),
    ("계획서", ["계획서", "계획", "work plan", "운영계획", "실시계획", "파견계획", "operation plan"]),
    ("방문/파견보고", ["방문", "파견결과", "courtesy", "working visit", "dispatch"]),
    ("워크숍", ["워크숍", "tlc", "workshop"]),
    ("회의/SCM", ["scm", "회의"]),
    ("행사/ICSE", ["icse", "컨퍼런스", "conference"]),
    ("진도/실적", ["progress", "term", "활동보고", "운영결과", "성과", "결과보고"]),
    ("보고서", ["보고서", "report"]),
]


def slug(s: str) -> str:
    s = re.sub(r"\.[a-zA-Z0-9]{2,5}$", "", s)          # 확장자 제거
    s = re.sub(r"[\(\)\[\]{}<>]", " ", s)
    s = re.sub(r"[^0-9A-Za-z가-힣]+", "-", s).strip("-")
    return s[:48] or "recipe"


def fmt_for(name: str) -> str:
    m = re.search(r"\.([a-zA-Z0-9]{2,5})\b", name)
    if m and m.group(1).lower() in EXT_FORMAT:
        return EXT_FORMAT[m.group(1).lower()]
    low = name.lower()
    if any(k in low for k in ("wrap up", "presentation", "ppt", "발표")):
        return "pptx"
    return "hwpx"   # 보고서류 기본값 (회사 HWPX 양식 적용 대상)


def category_for(name: str) -> str:
    low = name.lower()
    for cat, keys in CATEGORY_RULES:
        if any(k in low for k in keys):
            return cat
    return "기타"


def accept_for(label: str) -> list:
    m = re.search(r"\.([a-zA-Z0-9]{2,5})\b", label)
    if m:
        ext = m.group(1).lower()
        base = {ext} if ext in ("docx", "hwpx", "pdf", "xlsx", "pptx") else set()
        base |= {"pdf"}  # PDF 보조자료 항상 허용
        return sorted(base) or ["docx", "hwpx", "pdf"]
    return ["docx", "hwpx", "pdf", "xlsx", "pptx"]


def build_prompt(name: str, inputs: list) -> str:
    return (
        f"당신은 AIM 해외사업부 해외사업의 문서 작성 전문가입니다.\n"
        f"아래 입력 자료를 근거로 '{name}' 문서를 한국어로 작성하세요.\n"
        f"- 사실은 입력 자료에 근거하고, 추측이 필요한 부분은 [확인 필요]로 표시합니다.\n"
        f"- 보고서체의 정확하고 간결한 공식 문체를 사용합니다.\n\n"
        f"[입력 자료]\n{{{{inputs}}}}\n\n"
        f"[추가 지시]\n{{{{instruction}}}}\n"
    )


def main():
    import yaml
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl 필요: pip install openpyxl"); sys.exit(1)

    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(xlsx):
        print(f"엑셀을 찾을 수 없습니다: {xlsx}"); sys.exit(1)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    created, skipped = 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            no = int(row[0])
        except (TypeError, ValueError):
            continue
        out_name = str(row[1] or "").strip()
        if not out_name:
            continue
        rid = f"{no:02d}-{slug(out_name)}"
        path = os.path.join(HERE, f"{rid}.yaml")
        if os.path.exists(path):
            skipped += 1
            continue

        inputs = []
        for j, cell in enumerate(row[2:6]):
            label = str(cell or "").strip()
            if not label:
                continue
            inputs.append({
                "key": f"input{chr(ord('A') + j)}",
                "label": label,
                "required": (len(inputs) == 0),
                "accept": accept_for(label),
            })

        fmt = fmt_for(out_name)
        recipe = {
            "id": rid,
            "name": re.sub(r"\s+", " ", out_name),
            "category": category_for(out_name),
            "description": f"엑셀 목록표 {no}번 항목. 입력 자료로 '{out_name}' 초안 생성.",
            "output": {
                "format": fmt,
                "filename": slug(out_name) + "_{date}",
            },
            "inputs": inputs,
            "prompt": build_prompt(out_name, inputs),
        }
        with open(path, "w", encoding="utf-8") as f:
            yaml.safe_dump(recipe, f, allow_unicode=True, sort_keys=False)
        created += 1
        print(f"  + {rid}.yaml  [{fmt}]  입력 {len(inputs)}개")

    wb.close()
    print(f"\n완료: 생성 {created}개 / 기존유지 {skipped}개")


if __name__ == "__main__":
    main()
