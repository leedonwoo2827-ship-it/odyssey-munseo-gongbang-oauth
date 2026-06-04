"""XLSX 생성 — 구조화 표(openpyxl).

payload(table 모드):
  {"title": "...", "columns": ["A","B"], "rows": [[...],[...]]}
Markdown 이 오면 파이프 표를 파싱해 시트로 변환.
"""
from __future__ import annotations

from typing import Any, Dict, List


def _coerce_table(payload: Any) -> Dict:
    if isinstance(payload, dict) and "rows" in payload:
        return payload
    # Markdown 파이프 표 파싱
    md = payload if isinstance(payload, str) else ""
    columns: List[str] = []
    rows: List[List[str]] = []
    title = ""
    for line in (md or "").splitlines():
        s = line.strip()
        if s.startswith("#") and not title:
            title = s.lstrip("# ").strip()
            continue
        if s.startswith("|") and s.endswith("|"):
            cells = [c.strip() for c in s.strip("|").split("|")]
            if set("".join(cells)) <= set("-: "):  # 구분선
                continue
            if not columns:
                columns = cells
            else:
                rows.append(cells)
    if not columns and not rows:
        columns = ["내용"]
        rows = [[ln] for ln in (md or "").splitlines() if ln.strip()]
    return {"title": title or "표", "columns": columns, "rows": rows}


def render(payload: Any, out_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    data = _coerce_table(payload)
    wb = Workbook()
    ws = wb.active
    ws.title = (data.get("title") or "Sheet1")[:31]

    cols = data.get("columns", [])
    if cols:
        ws.append(cols)
        header_fill = PatternFill("solid", fgColor="1F2A44")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
    for row in data.get("rows", []):
        ws.append(list(row))

    # 열 너비 자동(대략)
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 60)

    wb.save(out_path)
    return out_path
